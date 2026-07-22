import difflib
import io
import re
from pathlib import Path

import fitz
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.config import settings
from app.services.path_service import _resolve_pdf_path

NUMBER_TOKEN_PATTERN = re.compile(r"\(?-?\d{1,3}(?:\.\d{3})+\)?|^-$")

# Nhan cac dong theo dung mau chuan NHNN (Thong tu 49/2014 + 27/2021) - giong
# nhau giua cac ngan hang vi la mau bat buoc, chi khac nhau ve so lieu. Danh
# sach nay KHONG kem so thu tu dau dong (I, 1, a, -...) vi phan do se bi bo
# qua khi so khop (xem _normalize_label).
B02A_LABELS = [
    "TÀI SẢN",
    "Tiền mặt, vàng bạc, đá quý",
    "Tiền gửi tại Ngân hàng Nhà nước",
    "Tiền gửi và cho vay các tổ chức tín dụng khác",
    "Tiền gửi tại các tổ chức tín dụng khác",
    "Cho vay các tổ chức tín dụng khác",
    "Dự phòng rủi ro",
    "Chứng khoán kinh doanh",
    "Chứng khoán kinh doanh",  # trung ten voi dong tren: 1 la muc cha (IV), 1 la muc con (IV.1)
    "Dự phòng rủi ro chứng khoán kinh doanh",
    "Các công cụ tài chính phái sinh và các tài sản tài chính khác",
    "Cho vay khách hàng",
    "Cho vay khách hàng",  # trung ten voi dong tren: 1 la muc cha (VI), 1 la muc con (VI.1)
    "Dự phòng rủi ro cho vay khách hàng",
    "Chứng khoán đầu tư",
    "Chứng khoán đầu tư sẵn sàng để bán",
    "Chứng khoán đầu tư giữ đến ngày đáo hạn",
    "Dự phòng rủi ro chứng khoán đầu tư",
    "Góp vốn, đầu tư dài hạn",
    "Đầu tư dài hạn khác",
    "Dự phòng giảm giá đầu tư dài hạn",
    "Tài sản cố định",
    "Tài sản cố định hữu hình",
    "Nguyên giá tài sản cố định",
    "Hao mòn tài sản cố định",
    "Tài sản cố định vô hình",
    "Nguyên giá tài sản cố định",  # trung ten: dung chung cho ca TSCD huu hinh va vo hinh
    "Hao mòn tài sản cố định",  # trung ten: dung chung cho ca TSCD huu hinh va vo hinh
    "Bất động sản đầu tư",
    "Nguyên giá bất động sản đầu tư",
    "Hao mòn bất động sản đầu tư",
    "Tài sản Có khác",
    "Các khoản phải thu",
    "Các khoản lãi, phí phải thu",
    "Tài sản thuế thu nhập doanh nghiệp hoãn lại",
    "Tài sản Có khác",  # trung ten voi dong o tren (XII): 1 la muc cha, 1 la muc con (XII.4)
    "Các khoản dự phòng rủi ro cho các tài sản Có nội bảng khác",
    "TỔNG TÀI SẢN",
    "NỢ PHẢI TRẢ VÀ VỐN CHỦ SỞ HỮU",
    "Các khoản nợ Chính phủ và Ngân hàng Nhà nước",
    "Tiền gửi và vay Chính phủ, Ngân hàng Nhà nước",
    "Giao dịch bán và mua lại trái phiếu Chính phủ với Kho bạc Nhà nước",
    "Tiền gửi và vay các tổ chức tín dụng khác",
    "Tiền gửi của các tổ chức tín dụng khác",
    "Vay các tổ chức tín dụng khác",
    "Tiền gửi của khách hàng",
    "Các công cụ tài chính phái sinh và các khoản nợ tài chính khác",
    "Vốn tài trợ, ủy thác đầu tư, cho vay tổ chức tín dụng chịu rủi ro",
    "Phát hành giấy tờ có giá",
    "Các khoản nợ khác",
    "Các khoản lãi, phí phải trả",
    "Các khoản phải trả và công nợ khác",
    "Dự phòng rủi ro khác",
    "TỔNG NỢ PHẢI TRẢ",
    "VỐN CHỦ SỞ HỮU",
    "Vốn của tổ chức tín dụng",
    "Vốn điều lệ",
    "Thặng dư vốn cổ phần",
    "Quỹ của tổ chức tín dụng",
    "Chênh lệch tỷ giá hối đoái",
    "Lợi nhuận chưa phân phối",
    "Lợi nhuận năm nay",
    "Lợi nhuận lũy kế năm trước",
    "TỔNG NỢ PHẢI TRẢ VÀ VỐN CHỦ SỞ HỮU",
    # Muc "CAC CHI TIEU NGOAI BAO CAO TINH HINH TAI CHINH" - 1 bang con nam
    # ngay sau bang can doi ke toan, van thuoc cung trang Mau B02a/TCTD-HN.
    "CÁC CHỈ TIÊU NGOÀI BÁO CÁO TÌNH HÌNH TÀI CHÍNH",
    "Bảo lãnh vay vốn",
    "Cam kết giao dịch hối đoái",
    "Cam kết mua ngoại tệ",
    "Cam kết bán ngoại tệ",
    "Cam kết giao dịch hoán đổi",
    "Cam kết trong nghiệp vụ L/C",
    "Bảo lãnh khác",
    "Các cam kết khác",
    "Lãi cho vay và phí phải thu chưa thu được",
    "Nợ khó đòi đã xử lý",
    "Tài sản và chứng từ khác",
]

B03A_LABELS = [
    "Thu nhập lãi và các khoản thu nhập tương tự",
    "Chi phí lãi và các chi phí tương tự",
    "Thu nhập lãi thuần",
    "Thu nhập từ hoạt động dịch vụ",
    "Chi phí hoạt động dịch vụ",
    "Lãi thuần từ hoạt động dịch vụ",
    "Lãi/(lỗ) thuần từ hoạt động kinh doanh ngoại hối",
    "Lãi/(lỗ) thuần từ mua bán chứng khoán kinh doanh",
    "Lãi/(lỗ) thuần từ mua bán chứng khoán đầu tư",
    "Thu nhập từ hoạt động khác",
    "Chi phí hoạt động khác",
    "Lãi thuần từ hoạt động khác",
    "Thu nhập từ góp vốn, mua cổ phần",
    "Chi phí hoạt động",
    "Lợi nhuận thuần từ hoạt động kinh doanh trước chi phí dự phòng rủi ro tín dụng",
    "Chi phí dự phòng rủi ro tín dụng",
    "Tổng lợi nhuận trước thuế",
    "Chi phí thuế thu nhập doanh nghiệp hiện hành",
    "Thu nhập thuế thu nhập doanh nghiệp hoãn lại",
    "Chi phí thuế thu nhập doanh nghiệp",
    "Lợi nhuận sau thuế",
    "Lãi cơ bản trên cổ phiếu",
]

B04A_LABELS = [
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG KINH DOANH",
    "Thu nhập lãi và các khoản thu nhập tương tự nhận được",
    "Chi phí lãi và các chi phí tương tự đã trả",
    "Thu nhập từ hoạt động dịch vụ nhận được",
    "Chênh lệch số tiền thực thu/(thực chi) từ hoạt động kinh doanh (ngoại tệ, vàng bạc, chứng khoán)",
    "Thu nhập/(chi phí) khác",
    "Tiền thu các khoản nợ đã được xử lý xóa, bù đắp bằng nguồn rủi ro",
    "Tiền chi trả cho nhân viên và hoạt động quản lý",
    "Tiền thuế thu nhập thực nộp trong kỳ",
    "Lưu chuyển tiền thuần từ hoạt động kinh doanh trước những thay đổi của tài sản và công nợ hoạt động",
    "(Tăng)/giảm các khoản tiền gửi và cho vay các tổ chức tín dụng khác",
    "(Tăng)/giảm các khoản về kinh doanh chứng khoán",
    "(Tăng)/giảm các công cụ tài chính phái sinh và các tài sản tài chính khác",
    "(Tăng)/giảm các khoản cho vay khách hàng",
    "Giảm nguồn dự phòng để xử lý rủi ro, xử lý, bù đắp tổn thất các khoản",
    "(Tăng)/giảm khác về tài sản hoạt động",
    "Tăng/(giảm) các khoản nợ Chính phủ và Ngân hàng Nhà nước",
    "Tăng/(giảm) tiền gửi và vay các tổ chức tín dụng khác",
    "Tăng/(giảm) tiền gửi của khách hàng",
    "Tăng/(giảm) phát hành giấy tờ có giá",
    "Tăng/(giảm) vốn tài trợ, ủy thác đầu tư, cho vay mà tổ chức tín dụng chịu rủi ro",
    "Tăng/(giảm) các công cụ tài chính phái sinh và các khoản nợ tài chính khác",
    "Tăng/(giảm) khác về công nợ hoạt động",
    "Chi từ các quỹ của tổ chức tín dụng",
    "LƯU CHUYỂN TIỀN THUẦN TỪ HOẠT ĐỘNG KINH DOANH",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ",
    "Mua sắm tài sản cố định",
    "Tiền thu từ thanh lý, nhượng bán tài sản cố định",
    "Mua sắm bất động sản đầu tư",
    "Tiền thu từ bán, thanh lý bất động sản đầu tư",
    "Tiền chi đầu tư, góp vốn vào các đơn vị khác",
    "Tiền thu đầu tư, góp vốn vào các đơn vị khác",
    "Tiền thu cổ tức và lợi nhuận được chia từ các khoản đầu tư, góp vốn dài hạn",
    "LƯU CHUYỂN TIỀN THUẦN TỪ HOẠT ĐỘNG ĐẦU TƯ",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH",
    "Tiền thu từ phát hành giấy tờ có giá dài hạn có đủ điều kiện tính vào vốn tự có và các khoản vốn vay dài hạn khác",
    "Tiền chi thanh toán giấy tờ có giá dài hạn có đủ điều kiện tính vào vốn tự có và các khoản vốn vay dài hạn khác",
    "Cổ tức trả cho cổ đông, lợi nhuận đã chia",
    "LƯU CHUYỂN TIỀN THUẦN TỪ HOẠT ĐỘNG TÀI CHÍNH",
    "LƯU CHUYỂN TIỀN THUẦN TRONG KỲ",
    "TIỀN VÀ CÁC KHOẢN TƯƠNG ĐƯƠNG TIỀN TẠI NGÀY 1 THÁNG 1",
    "Điều chỉnh ảnh hưởng của thay đổi tỷ giá",
    "TIỀN VÀ CÁC KHOẢN TƯƠNG ĐƯƠNG TIỀN TẠI NGÀY 31 THÁNG 3",
    # Muc liet ke chi tiet "Tien va cac khoan tuong duong tien gom co" o cuoi
    # trang - can co trong schema de khong bi so khop nham sang cac dong khac
    # co ten gan giong (vi du "Tang/(giam) tien gui va vay cac to chuc tin
    # dung khac").
    "Tiền mặt, vàng bạc, đá quý",
    "Tiền gửi thanh toán tại Ngân hàng Nhà nước",
    "Tiền gửi tại các tổ chức tín dụng khác",
    "Chứng khoán đầu tư",
]

# marker: chuoi dac trung de nhan biet trang nao thuoc mau nao (in san goc
# tren trang, xem _detect_template).
TEMPLATES = {
    "B02a": {"marker": "B02a/TCTD-HN", "labels": B02A_LABELS},
    "B03a": {"marker": "B03a/TCTD-HN", "labels": B03A_LABELS},
    "B04a": {"marker": "B04a/TCTD-HN", "labels": B04A_LABELS},
}

# Mau B05a (thuyet minh) co hang chuc de muc khac nhau, tuy bao cao, khong the
# liet ke het thanh 1 danh sach nhan co dinh nhu 3 mau tren -> xu ly rieng:
# trich nguyen trang, co bang thi lay bang, khong co bang thi lay van ban.
B05A_MARKER = "B05a/TCTD-HN"
B05A_SKIP_PREFIXES = (
    "NGÂN HÀNG",
    "Mẫu",
    "THUYẾT MINH BÁO CÁO",
    "QUÝ I NĂM",
    "BÁO CÁO TÀI CHÍNH",
)

TEMPLATE_TITLES = {
    "B02a": "BÁO CÁO TÌNH HÌNH TÀI CHÍNH HỢP NHẤT GIỮA NIÊN ĐỘ (Mẫu B02a/TCTD-HN)",
    "B03a": "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG HỢP NHẤT GIỮA NIÊN ĐỘ (Mẫu B03a/TCTD-HN)",
    "B04a": "BÁO CÁO LƯU CHUYỂN TIỀN TỆ HỢP NHẤT GIỮA NIÊN ĐỘ (Mẫu B04a/TCTD-HN)",
}

# Ten cot gia tri - de trung tinh (khong ghi chet ngay thang cu the) vi ky bao
# cao thay doi tuy file. So cot phai khop voi so cot gia tri thuc te cua mau.
TEMPLATE_VALUE_HEADERS = {
    "B02a": ["Chỉ tiêu", "Kỳ báo cáo", "Kỳ trước"],
    "B03a": [
        "Chỉ tiêu",
        "Quý này - Kỳ báo cáo",
        "Quý này - Kỳ trước",
        "Luỹ kế - Kỳ báo cáo",
        "Luỹ kế - Kỳ trước",
    ],
    "B04a": ["Chỉ tiêu", "Kỳ báo cáo", "Kỳ trước"],
}

LABEL_PREFIX_PATTERN = re.compile(r"^(?:[IVXLCDM]+|\d+|[a-z])\s+|^-\s*", re.IGNORECASE)


def _page_has_marker(page, marker: str) -> bool:
    # Chi coi la trang THUOC ve mau do neu marker xuat hien nhu 1 dong RIENG,
    # NGAN o dau trang (giong cach "Mau B0Xa/TCTD-HN" luon duoc in rieng 1
    # dong ngay duoi ten ngan hang tren moi trang that). Trang muc luc cung
    # nhac den ten cac mau, nhung nam giua cau van dai - can loai truong hop
    # nay, khong thi tat ca cac trang muc luc/tom tat se bi nhan nham.
    lines = [line.strip() for line in page.get_text().split("\n") if line.strip()]
    for line in lines[:10]:
        if marker in line and len(line) <= len(marker) + 10:
            return True
    return False


def _detect_template(page) -> str | None:
    for name, info in TEMPLATES.items():
        if _page_has_marker(page, info["marker"]):
            return name
    return None


def _normalize_label(text: str) -> str:
    # Bo ky hieu danh muc dau dong (so La Ma, so thuong, chu thuong, dau '-')
    # de so sanh cong bang voi danh sach nhan trong schema (khong kem ky hieu).
    text = LABEL_PREFIX_PATTERN.sub("", text.strip())
    return re.sub(r"\s+", " ", text).lower().strip()


def _match_schema_label(extracted_label: str, schema_labels: list[str], threshold: float = 0.6) -> str | None:
    normalized = _normalize_label(extracted_label)
    if not normalized:
        return None

    best_label, best_score = None, 0.0
    for label in schema_labels:
        score = difflib.SequenceMatcher(None, normalized, label.lower()).ratio()
        if score > best_score:
            best_label, best_score = label, score

    return best_label if best_score >= threshold else None


def _find_value_column_start(rows: list) -> int:
    # Cot "tien" duoc xac dinh 1 lan cho CA BANG: la cot nho nhat ma co it nhat
    # 1 hang bat ky trong bang CHUA mot con so dinh dang dung (vi du "17.220.377"),
    # khong doi hoi ca o phai la so 100% - vi co bang find_tables() gop nham chu
    # con lai cua tu bi cat ngang vao chung o voi con so.
    max_cols = max((len(row) for row in rows), default=0)
    for col in range(max_cols):
        for row in rows:
            if col < len(row) and NUMBER_TOKEN_PATTERN.search((row[col] or "").strip()):
                return col
    return max_cols


def _split_multi_number_cell(cell: str) -> list[str]:
    # find_tables() co the gop 2 con so (2 nam) vao chung 1 o do luoi cot qua
    # tho. Neu 1 o co tu 2 con so dinh dang tro len, tach thanh nhieu o rieng.
    # Neu chi co 1 (hoac 0) con so, giu nguyen ca o (vi du "VII.2 83.036" giu
    # nguyen ma thuyet minh di kem, khong tach nham).
    tokens = [m.group() for m in NUMBER_TOKEN_PATTERN.finditer(cell)]
    return tokens if len(tokens) >= 2 else [cell]


def _merge_split_words(page, table, row_index: int, cells: list, value_start: int) -> list:
    cell_boxes = table.rows[row_index].cells[value_start:]
    value_cells = cells[value_start:]
    if not cell_boxes or not any(value_cells):
        # Hang von di rong (theo find_tables()) thi bo qua, khong doc lai gi ca.
        # Neu van goi get_text() tren vung toa do nay, co the "hut" nham chu tu
        # dong ke sat ben canh (do cac dong tieu de nam qua sat nhau).
        return value_cells

    row_bbox = (
        min(b[0] for b in cell_boxes),
        min(b[1] for b in cell_boxes),
        max(b[2] for b in cell_boxes),
        max(b[3] for b in cell_boxes),
    )
    words = page.get_text("words", clip=row_bbox)
    # cac dong sat nhau khong co khoang cach (y1 dong tren = y0 dong duoi), nen
    # clip co the "an" ca chu cua dong ke ben. Chi giu tu co diem giua theo
    # truc y nam dung trong pham vi dong nay, tranh dinh chu tu dong ben canh.
    row_y0, row_y1 = row_bbox[1], row_bbox[3]
    words = [w for w in words if row_y0 <= (w[1] + w[3]) / 2 < row_y1]

    # neu 1 tu that trong PDF de len ranh gioi giua 2 o mot doan DANG KE (khong
    # phai chi le vai diem anh do bo tron font), nghia la find_tables() da cat
    # ngang tu do -> can gop 2 o lai. Nguong toi thieu de loai truong hop mot
    # con so nam sat vien nhung van thuoc tron 1 o (chi lan qua 1-2pt).
    MIN_OVERLAP = 2.0
    merge_with_next = [False] * len(cell_boxes)
    for word in words:
        word_x0, word_x1 = word[0], word[2]
        touched = [
            i for i, box in enumerate(cell_boxes)
            if min(word_x1, box[2]) - max(word_x0, box[0]) > MIN_OVERLAP
        ]
        if len(touched) > 1:
            for i in range(min(touched), max(touched)):
                merge_with_next[i] = True

    result: list[str] = []
    i = 0
    while i < len(cell_boxes):
        j = i
        while j < len(cell_boxes) - 1 and merge_with_next[j]:
            j += 1

        if j == i:
            result.append(value_cells[i])
        else:
            x0 = min(cell_boxes[k][0] for k in range(i, j + 1))
            y0 = min(cell_boxes[k][1] for k in range(i, j + 1))
            x1 = max(cell_boxes[k][2] for k in range(i, j + 1))
            y1 = max(cell_boxes[k][3] for k in range(i, j + 1))
            text = page.get_text(clip=(x0, y0, x1, y1))
            result.append(re.sub(r"\s+", " ", text).strip())
            result.extend([""] * (j - i))

        i = j + 1

    return result


def _is_row_label_bold(page, table, row_index: int, row: list, value_start: int) -> bool:
    # Doc thang thong tin font tu PDF (flag bit 16 = in dam) thay vi doan qua
    # chu HOA - vi cac dong "muc cha" (I, II, III, IV...) trong file nay in
    # dam nhung KHONG viet hoa toan bo, khac voi cac dong tong ("TÀI SẢN")
    # vua in dam vua viet hoa. Day la cach phan biet dung muc cha/muc con
    # trung ten (vi du 2 dong "Chứng khoán kinh doanh").
    cells = [(cell or "").strip() for cell in row]
    label_cell_indices = [i for i in range(min(value_start, len(cells))) if cells[i]]
    if not label_cell_indices:
        return False

    cell_boxes = table.rows[row_index].cells
    x0 = min(cell_boxes[i][0] for i in label_cell_indices)
    y0 = min(cell_boxes[i][1] for i in label_cell_indices)
    x1 = max(cell_boxes[i][2] for i in label_cell_indices)
    y1 = max(cell_boxes[i][3] for i in label_cell_indices)

    text_dict = page.get_text("dict", clip=(x0, y0, x1, y1))
    for block in text_dict.get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                if span.get("flags", 0) & 2**4 or "bold" in span.get("font", "").lower():
                    return True
    return False


def _build_excel_row(page, table, row_index: int, row: list, value_start: int) -> list:
    cells = [(cell or "").strip() for cell in row]
    merged_cells = _merge_split_words(page, table, row_index, cells, value_start)

    value_cells: list[str] = []
    for cell in merged_cells:
        value_cells.extend(_split_multi_number_cell(cell))

    label_cell_indices = [i for i in range(min(value_start, len(cells))) if cells[i]]
    if not label_cell_indices:
        return ["", *value_cells]

    # find_tables() cat text theo luoi cot co dinh, co the cat ngang giua tu.
    # Nen khong noi cac manh da bi cat, ma lay lai vung toa do (bbox) cua
    # cac o-label roi doc lai text goc bang get_text() cho dung khoang cach.
    cell_boxes = table.rows[row_index].cells
    x0 = min(cell_boxes[i][0] for i in label_cell_indices)
    y0 = min(cell_boxes[i][1] for i in label_cell_indices)
    x1 = max(cell_boxes[i][2] for i in label_cell_indices)
    y1 = max(cell_boxes[i][3] for i in label_cell_indices)

    label = page.get_text(clip=(x0, y0, x1, y1))
    label = re.sub(r"\s+", " ", label).strip()

    return [label, *value_cells]


def _is_clean_data_row(row: list[str]) -> bool:
    # B05a khong co schema de doi chieu, nen khong the biet chac 1 hang co
    # dung khong. Chi tin duoc hang dang "nhan + 2 hoac 3 so sach" (moi o gia
    # tri la MOT con so dinh dang dung, khong con lan chu) - day la dang pho
    # bien nhat trong cac thuyet minh dang bang. Hang khong khop dang nay
    # (tieu de nhieu tang, o bi dinh chu...) thi KHONG co ich khi ep vao dung
    # cot, nen se duoc gop lai thanh 1 dong van ban o cho goi (xem duoi).
    label = (row[0] or "").strip()
    values = [(v or "").strip() for v in row[1:]]
    non_empty_values = [v for v in values if v]
    if not label or not (2 <= len(non_empty_values) <= 3):
        return False
    return all(NUMBER_TOKEN_PATTERN.fullmatch(v) for v in non_empty_values)


def _extract_b05a_block(page) -> list[list[str]]:
    # Neu trang co bang: trich giong het cach 1 (dung lai toan bo logic tach
    # so da co) - khong doi chieu schema vi B05a khong co danh sach nhan co
    # dinh. Neu trang khong co bang (thuyet minh dang van ban thuong): tra ve
    # tung dong van ban, bo qua cac dong tieu de lap lai o moi trang.
    tables = page.find_tables(strategy="text").tables
    if tables:
        block_rows: list[list[str]] = []
        for table in tables:
            rows = table.extract()
            if not rows:
                continue
            value_start = _find_value_column_start(rows)
            for row_index, row in enumerate(rows):
                built = _build_excel_row(page, table, row_index, row, value_start)
                if not any((cell or "").strip() for cell in built):
                    continue
                if _is_clean_data_row(built):
                    block_rows.append(built)
                else:
                    # Hang khong ro rang (tieu de nhieu tang, bang long nhau...)
                    # -> gop het thanh 1 dong van ban thay vi ep sai cot.
                    text = " ".join(cell for cell in built if (cell or "").strip())
                    block_rows.append([text])
        return block_rows

    lines = [line.strip() for line in page.get_text().split("\n") if line.strip()]
    filtered = [
        line for line in lines
        if not line.isdigit() and not line.startswith(B05A_SKIP_PREFIXES)
    ]
    return [[line] for line in filtered]


def _detect_printed_page_number(page) -> str | None:
    # So trang IN SAN tren chinh tai lieu thuong nam rieng 1 dong, ngay sau
    # ten ngan hang, o dau trang - khac voi vi tri trang trong file PDF (file
    # co bia + muc luc o dau nen bi lech). Doc vai dong dau, tim dong chi toan
    # chu so.
    text = page.get_text()
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    for line in lines[:5]:
        if line.isdigit():
            return line
    return None


def export_pdf_tables_to_excel(document_id: str) -> bytes:
    pdf_path = _resolve_pdf_path(document_id)

    workbook = Workbook()
    default_sheet = workbook.active
    used_default_sheet = False
    used_titles: set[str] = set()

    try:
        with fitz.open(str(pdf_path)) as pdf:
            for page_index, page in enumerate(pdf):
                tables = page.find_tables(strategy="text").tables
                if not tables: continue

                printed_number = _detect_printed_page_number(page)
                preferred_title = f"Trang {printed_number}" if printed_number else None
                # Ten du phong dung tien to KHAC HAN, de khong bao gio trung
                # voi ten dat theo so trang in - tranh hieu ung day chuyen khi
                # 1 trang du phong vo tinh trung voi so trang in that cua 1
                # trang khac.
                if preferred_title and preferred_title not in used_titles:
                    sheet_title = preferred_title
                else:
                    sheet_title = f"Trang (file) {page_index + 1}"
                used_titles.add(sheet_title)

                if not used_default_sheet:
                    sheet = default_sheet
                    sheet.title = sheet_title
                    used_default_sheet = True
                else:
                    sheet = workbook.create_sheet(title=sheet_title)

                for table in tables:
                    rows = table.extract()
                    if not rows: continue
                    value_start = _find_value_column_start(rows)
                    for row_index, row in enumerate(rows):
                        sheet.append(_build_excel_row(page, table, row_index, row, value_start))

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Cannot export tables") from exc

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_pdf_by_template_to_excel(document_id: str) -> bytes:
    # Cach 2: thay vi doan luoi cot tu find_tables(), doi chieu tung dong voi
    # danh sach nhan CHUAN cua mau (B02a/B03a/B04a). Van tai su dung dung logic
    # tach so da co (_build_excel_row) de lay gia tri - chi khac o cho ket qua
    # duoc LOC va SAP XEP lai theo dung thu tu chuan cua mau, khong con dong
    # rac (tieu de, dong trong) va khong con phu thuoc luoi cot doan sai.
    pdf_path = _resolve_pdf_path(document_id)

    # Gom du lieu theo tung mau, TRAI DAI qua nhieu trang (vi du B02a nam o
    # ca trang 3 va trang 4) - de cuoi cung xuat 1 sheet duy nhat cho 1 mau.
    # Moi nhan luu 1 HANG DOI cac gia tri (khong phai 1 gia tri duy nhat), vi
    # file goc co the dung LAI CUNG 1 TEN cho ca muc cha va muc con (vi du
    # "IV Chung khoan kinh doanh" va "1 Chung khoan kinh doanh" cung ten) -
    # neu chi luu 1 gia tri thi dong sau se ghi de mat dong truoc. Khi xuat,
    # tieu thu hang doi theo dung thu tu (xem schema co khai bao trung ten
    # may lan, xem cac danh sach B0XA_LABELS).
    matched_by_template: dict[str, dict[str, list[list]]] = {name: {} for name in TEMPLATES}
    # B05a: gom theo tung trang (khong doi chieu schema), moi phan tu la
    # (so_trang_in, cac_hang_cua_trang_do).
    b05a_pages: list[tuple[str, list[list[str]]]] = []
    found_any_page = False

    try:
        with fitz.open(str(pdf_path)) as pdf:
            for page in pdf:
                if _page_has_marker(page, B05A_MARKER):
                    block_rows = _extract_b05a_block(page)
                    if block_rows:
                        printed_number = _detect_printed_page_number(page) or "?"
                        b05a_pages.append((printed_number, block_rows))
                        found_any_page = True
                    continue

                template_name = _detect_template(page)
                if not template_name:
                    continue

                tables = page.find_tables(strategy="text").tables
                if not tables:
                    continue

                schema_labels = TEMPLATES[template_name]["labels"]
                matched_rows = matched_by_template[template_name]

                for table in tables:
                    rows = table.extract()
                    if not rows:
                        continue
                    value_start = _find_value_column_start(rows)

                    # Nhan bi PDF ngat xuong dong (vi du "1 Thu nhap lai va cac
                    # khoan thu nhap" / "tuong tu VI.1  17.220.377 ...") bi tach
                    # thanh 2 hang rieng khi trich: hang dau (chi co nhan, chua
                    # co so) va hang sau (co so nhung nhan cut, khop kem). Nho
                    # lai muc schema vua khop o hang "chi co nhan" gan nhat, de
                    # neu hang KE TIEP co so nhung tu no khong khop duoc gi thi
                    # gan so lieu do cho dung muc do.
                    pending_entry: list | None = None  # [values, is_bold], mutate tai cho
                    pending_row_index: int | None = None

                    for row_index, row in enumerate(rows):
                        built = _build_excel_row(page, table, row_index, row, value_start)
                        label, values = built[0], built[1:]
                        if not label:
                            continue

                        has_values = any((v or "").strip() for v in values)
                        matched_label = _match_schema_label(label, schema_labels)

                        if not has_values:
                            if matched_label:
                                is_bold = _is_row_label_bold(page, table, row_index, row, value_start)
                                entry = [list(values), is_bold]
                                matched_rows.setdefault(matched_label, []).append(entry)
                                pending_entry = entry
                                pending_row_index = row_index
                                found_any_page = True
                            continue

                        # Uu tien noi vao "o" dang cho (neu ke ngay sau hang
                        # chi-co-nhan) TRUOC khi xet ban than dong nay co tu
                        # khop duoc gi khong - vi phan nhan cut con lai sau khi
                        # bi ngat dong (vi du "tai chinh khac V.3") de tu khop
                        # NHAM sang 1 muc khac chi vi trung vai chu ngau nhien.
                        if pending_entry is not None and row_index == pending_row_index + 1:
                            pending_entry[0] = values
                            found_any_page = True
                        elif matched_label:
                            is_bold = _is_row_label_bold(page, table, row_index, row, value_start)
                            matched_rows.setdefault(matched_label, []).append([values, is_bold])
                            found_any_page = True

                        pending_entry = None

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Cannot export by template") from exc

    if not found_any_page:
        raise HTTPException(status_code=400, detail="Khong nhan dien duoc mau bao cao nao trong file nay")

    workbook = Workbook()
    default_sheet = workbook.active
    used_default_sheet = False

    for template_name, matched_rows in matched_by_template.items():
        if not matched_rows:
            continue

        if not used_default_sheet:
            sheet = default_sheet
            sheet.title = template_name
            used_default_sheet = True
        else:
            sheet = workbook.create_sheet(title=template_name)

        header_row = TEMPLATE_VALUE_HEADERS.get(template_name, ["Chỉ tiêu"])

        # Dong tieu de bao cao, gop o cho de doc, in dam + co lon hon.
        sheet.append([TEMPLATE_TITLES.get(template_name, template_name)])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_row))
        sheet.cell(row=1, column=1).font = Font(bold=True, size=13)

        sheet.append([])

        # Dong ten cot, in dam.
        sheet.append(header_row)
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

        # Ghi theo DUNG THU TU chuan cua mau (khong theo thu tu tim thay tren
        # trang), dong nao khong tim thay so lieu tuong ung thi bo qua. Neu 1
        # ten xuat hien nhieu lan trong schema (mau cha/con trung ten), moi
        # lan se lay 1 gia tri KE TIEP trong hang doi (dung thu tu da gap).
        consumed_count: dict[str, int] = {}
        for schema_label in TEMPLATES[template_name]["labels"]:
            entries = matched_rows.get(schema_label)
            if not entries:
                continue
            next_index = consumed_count.get(schema_label, 0)
            if next_index >= len(entries):
                continue
            consumed_count[schema_label] = next_index + 1
            values, is_bold = entries[next_index]

            sheet.append([schema_label, *values])
            # In dam neu: (a) ten viet HOA toan bo (vi du "TÀI SẢN", "TỔNG TÀI
            # SẢN"), hoac (b) chinh PDF goc da in dam dong nay (doc tu font
            # that, xem _is_row_label_bold) - day la cach phan biet dung
            # muc cha/muc con trung ten (vi du 2 dong "Chứng khoán kinh doanh":
            # dong cha in dam, dong con thi khong).
            if schema_label.isupper() or is_bold:
                for cell in sheet[sheet.max_row]:
                    cell.font = Font(bold=True)

        sheet.column_dimensions["A"].width = 55
        for col_letter in "BCDEF":
            sheet.column_dimensions[col_letter].width = 20

    if b05a_pages:
        if not used_default_sheet:
            sheet = default_sheet
            sheet.title = "B05a"
            used_default_sheet = True
        else:
            sheet = workbook.create_sheet(title="B05a")

        sheet.append(["THUYẾT MINH BÁO CÁO TÀI CHÍNH CHỌN LỌC HỢP NHẤT (Mẫu B05a/TCTD-HN)"])
        sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
        sheet.append([])

        # Khong co danh sach nhan chuan nhu 3 mau kia, nen chi gan tieu de
        # "Trang X" (so trang in tren tai lieu) truoc noi dung cua trang do,
        # giu nguyen thu tu xuat hien trong file - de nguoi doc con doi chieu
        # duoc voi ban PDF goc.
        for printed_number, rows in b05a_pages:
            sheet.append([f"--- Trang {printed_number} ---"])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
            for row in rows:
                sheet.append(row)
            sheet.append([])

        sheet.column_dimensions["A"].width = 70
        for col_letter in "BCDEF":
            sheet.column_dimensions[col_letter].width = 20

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()